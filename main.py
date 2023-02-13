import os
import time
import json
import random
import requests
import sys

import pyrogram
from pyrogram import Client
from PIL import Image

import configparser
import openpyxl

config = configparser.ConfigParser()
config.read('config.ini')

API_ID = config.get('Telegram', 'api_id')
API_HASH = config.get('Telegram', 'api_hash')
phone_number = config.get('Telegram', 'phone_number')

with open('config.ini', 'w') as configfile:
    config.write(configfile)


client = pyrogram.Client("Sessions/", API_ID, API_HASH, phone_number=phone_number)
client.start()


class Menu:
    def __init__(self):
        self.actions = {
            1: self.add_users_by_link,
            2: self.add_users_by_excel,
            3: self.add_groups_by_excel,
            4: self.export_users_from_group,
            5: self.extract_ids,
            6: self.warmup_profile,
            7: self.update_telegram_credentials,
        }

    def run(self):
        while True:
            print("LEAD ROBOT - TELEGRAM")
            print("\n1- Add through link")
            print("2- Add through Excel")
            print("3- Export groups from your number")
            print("4- Export users from a group")
            print("5- Extract IDs")
            print("6- Warm up profile")
            print("7- Update Telegram credentials")

            action = int(input("Choose an option: "))

            if action in self.actions:
                self.actions[action]()
            else:
                print("Invalid option.")


    def add_users_by_link(self):

        source_group_link = input("Link of the source group: ")
        destination_group_link = input("Link of the destination group: ")

        source_group = client.get_chat(source_group_link)
        source_group_id = source_group.id

        destination_group = client.get_chat(destination_group_link)
        destination_group_id = destination_group.id

        members = list(client.get_chat_members(source_group_id))
        random.shuffle(members)

        num_users = int(input("Enter the number of users you want to add: "))

        sleep_time = int(input("Enter the time in seconds to add each member: "))

        added_count = 0

        try:
            with open("sent_messages.json", "r") as f:
                sent_messages = json.load(f)
        except FileNotFoundError:
            sent_messages = []

        for i, member in enumerate(members):
            if added_count >= num_users:
                with open("sent_messages.json", "w") as f:
                    json.dump(sent_messages, f)
                print("Users successfully added")
                break

            time.sleep(1)
            try:
                user_id = member.user.id
                if user_id not in sent_messages:
                    client.add_chat_members(destination_group_id, user_id)
                    print(f"Usuário {member.user.first_name} Added successfully!")
                    added_count += 1
                    print(f"Extraindo {added_count}/{num_users}")
                    sent_messages.append(user_id)
                    time.sleep(sleep_time)
            except Exception as e:
                print(f"Error adding user with username {member}: {e}")
                continue

    def add_users_by_excel(self):
        destination_group_link = input("Group destination link: ")

        destination_group = client.get_chat(destination_group_link)
        destination_group_id = destination_group.id

        filename = input("Name of the Excel file: ")

        workbook = openpyxl.load_workbook(f"{filename}.xlsx")
        sheet = workbook.active

        column = int(input("Enter the column: "))
        min_row = int(input("Enter the starting row: "))

        usernames = [
            row[column - 1]
            for row in sheet.iter_rows(min_row=min_row, values_only=True)
        ]
        random.shuffle(usernames)

        num_users = int(input("Enter the number of users you want to add: "))

        sleep_time = int(input("Enter the time in seconds to add each member: "))

        added_count = 0

        try:
            with open("sent_messages.json", "r") as f:
                sent_messages = json.load(f)
        except FileNotFoundError:
            sent_messages = []

        for i, username in enumerate(usernames):
            if added_count >= num_users:
                with open("sent_messages.json", "w") as f:
                    json.dump(sent_messages, f)
                print("Users added successfully")
                break

            if username not in sent_messages:
                time.sleep(1)
                try:
                    client.add_chat_members(destination_group_id, username)
                    print(f"User {username} added successfully!")
                    added_count += 1
                    print(f"Extracting {added_count}/{num_users}")
                    sent_messages.append(username)
                    time.sleep(sleep_time)
                except Exception as e:
                    print(f"Error adding user with username {username}: {e}")
                    continue

    def add_groups_by_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet["A1"] = "Group Name"
        sheet["B1"] = "ID"

        chats = client.get_dialogs()

        for row_num, chat in enumerate(chats, start=2):
            sheet.cell(row=row_num, column=1).value = chat.chat.title
            sheet.cell(row=row_num, column=2).value = chat.chat.id
        workbook.save("Grupos.xlsx")

    def export_users_from_group(self):

        print("1 - Export users from a group via link")
        print("2 - Export users from a group via ID")

        action = int(input("Choose an option: "))

        if action == 1:
            group_link = input("Enter Telegram Link: ")
            chat_id = client.get_chat(group_link).id

        elif action == 2:
            chat_id = int(input("Enter Telegram ID: "))

        members = client.get_chat_members(chat_id)

        admins = []

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet["A1"] = "Username"
        sheet["B1"] = "ID"
        sheet["C1"] = "Nome"

        print("Exporting users...")
        row_num = 2
        for member in members:
            if member.user.username is not None:
                sheet.cell(row=row_num, column=1).value = member.user.username
                sheet.cell(row=row_num, column=2).value = member.user.id
                sheet.cell(row=row_num, column=3).value = member.user.first_name
                row_num += 1

        name_excel = input("Name of excel: ")
        workbook.save(f"{name_excel}.xlsx")

    def extract_ids(self):
        source_group_link = input("Source group link: ")

        source_group = client.get_chat(source_group_link)
        source_group_id = source_group.id

        members = client.get_chat_members(source_group_id)

        try:
            with open("sent_messages.json", "r") as f:
                sent_messages = json.load(f)
        except FileNotFoundError:
            sent_messages = []

        for member in members:
            user_id = member.user.id
            if user_id not in sent_messages:
                sent_messages.append(user_id)

        with open("sent_messages.json", "w") as f:
            json.dump(sent_messages, f)

        print("Users added successfully")
        print(f"{len(sent_messages)} Users successfully extracted")

    def warmup_profile(self):

        def message():
            usernames = ["@user1", "@user2", "@user3", "@user4", "@user5"]

            for username in usernames:
                message = f"Hi {username}, How are You?"
                try:
                    client.send_message(username, message)
                    print(f"Message sent to {username}")
                    time.sleep(75)
                except Exception as e:
                    print(f"Unable to send message to {username} due to error: {e}")

        print("Do you want to add user, bio and photo? [Y/N]")
        action = input("Choose an option: ").upper()

        if action == "S":
            response = requests.get("https://randomuser.me/api/")
            data = response.json()
            first_name = data['results'][0]['name']['first']
            last_name = data['results'][0]['name']['last']
            name = f"{first_name} {last_name}"
            photo = data['results'][0]['picture']['large']
            bio = "Hi, How are You?"

            response = requests.get(photo)
            path = "path/to/save"
            if not os.path.exists(path):
                os.makedirs(path)
            open(f"{path}/photo.jpg", "wb").write(response.content)
            with Image.open(f"{path}/photo.jpg") as im:
                width, height = im.size
                if width < 200 or height < 200:
                    size = (200, 200)
                    im = im.resize(size)
                    im.save(f"{path}/photo.jpg")
            client.set_profile_photo(photo=f"{path}/photo.jpg")
            time.sleep(45)
            client.update_profile(first_name=name)
            time.sleep(52)
            client.update_profile(bio=bio)
            print("Photo, name, and bio added successfully")
            time.sleep(78)
            message()

        elif action == "N":
            message()

    def update_telegram_credentials(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        config.set('Telegram', 'api_id', input("Enter the new API_ID: "))
        config.set('Telegram', 'api_hash', input("Enter the new API_HASH: "))
        config.set('Telegram', 'phone_number', input("Enter the new PHONE: "))
        with open('config.ini', 'w') as configfile:
            config.write(configfile)

if __name__ == '__main__':
    menu = Menu()
    menu.run()